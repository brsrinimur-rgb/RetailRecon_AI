"""
REGRESSION_POS_GL_V44_STATUS_FORMATTING_TOTALS.py

Three small, presentation-focused improvements requested after the V42.2
Summary redesign was verified working in production:

  1. gl_summary now carries the same tolerance-based Status column as
     store_summary (logic/pos_gl_reconciliation.py).
  2. Both Summary tables get OK/REVIEW conditional formatting in the
     Excel export (pages/35_POS_GL_Reconciliation.py).
  3. A bold "TOTAL" row is written below each table, showing the same
     POS/GL/Difference totals already proven to tie out internally --
     now visible in the workbook itself.

None of this touches Store+Date bucketing, matching, duplicate
detection, or tolerance logic -- purely additive to the two summary
DataFrames and the Excel writer.

Run: python3 REGRESSION_POS_GL_V44_STATUS_FORMATTING_TOTALS.py
"""
from pathlib import Path
import sys
import io
import pandas as pd

root = Path(__file__).resolve().parent
sys.path.insert(0, str(root))
from logic import pos_gl_reconciliation as pgl

# ---------------------------------------------------------------------
# 1. gl_summary Status column, same tolerance rule as store_summary.
# ---------------------------------------------------------------------
pos = pd.DataFrame([
    {"store_code": "601", "pos_date": "2026-07-01", "pos_amount": 1000.00, "provider": "MADA",
     "reference": "A1", "auth_code": "A1", "source_row": 1, "merchant_id": "M1"},
    {"store_code": "602", "pos_date": "2026-07-01", "pos_amount": 500.00, "provider": "MADA",
     "reference": "A2", "auth_code": "A2", "source_row": 2, "merchant_id": "M1"},
    {"store_code": "603", "pos_date": "2026-07-01", "pos_amount": 2000.00, "provider": "MADA",
     "reference": "A3", "auth_code": "A3", "source_row": 3, "merchant_id": "M1"},
])
gl = pd.DataFrame([
    {"store_code": "601", "gl_date": "2026-07-01", "gl_signed_amount": 1000.00, "gl_amount": 1000.00,
     "main_account": "11020907", "voucher": "V1", "journal": "J1", "source_row": 1, "source_file": "gl.xlsx"},
    {"store_code": "602", "gl_date": "2026-07-01", "gl_signed_amount": 500.01, "gl_amount": 500.01,
     "main_account": "11020907", "voucher": "V2", "journal": "J2", "source_row": 2, "source_file": "gl.xlsx"},
    {"store_code": "603", "gl_date": "2026-07-01", "gl_signed_amount": 1500.00, "gl_amount": 1500.00,
     "main_account": "11020907", "voucher": "V3", "journal": "J3", "source_row": 3, "source_file": "gl.xlsx"},
])
r = pgl.reconcile_pos_to_gl_by_bucket(pos, gl, tolerance=0.50, max_bucket_tolerance=25.00)

assert "Status" in r["gl_summary"].columns
assert r["gl_summary"]["Status"].iloc[0] == "REVIEW", (
    "combined 11020907 difference (499.99) exceeds SAR 25 tolerance -- must be REVIEW"
)
print("[PASS] gl_summary carries a Status column using the same SAR 25 tolerance rule "
      "as store_summary.")

# Empty case: Status column must still exist, just empty.
empty_pos = pd.DataFrame(columns=["store_code", "pos_date", "pos_amount", "provider",
                                   "reference", "auth_code", "source_row", "merchant_id"])
empty_gl = pd.DataFrame(columns=["store_code", "gl_date", "gl_signed_amount", "gl_amount",
                                  "main_account", "voucher", "journal", "source_row", "source_file"])
r_empty = pgl.reconcile_pos_to_gl_by_bucket(empty_pos, empty_gl)
assert "Status" in r_empty["gl_summary"].columns
assert r_empty["gl_summary"].empty
print("[PASS] gl_summary Status column present even when the result is empty -- no crash.")


# ---------------------------------------------------------------------
# 2 & 3. Excel writer: conditional formatting + TOTAL row, exercised by
# running the actual writer logic against real function output and
# inspecting the produced workbook with openpyxl.
# ---------------------------------------------------------------------
def _write_summary_sheet(r):
    """Mirrors the Summary-sheet block in pages/35_POS_GL_Reconciliation.py."""
    b = io.BytesIO()
    with pd.ExcelWriter(b, engine="openpyxl") as w:
        _store_summary = r.get("store_summary", pd.DataFrame())
        _gl_summary = r.get("gl_summary", pd.DataFrame())
        # V42.5: "Mapping Required" is a raw flag, not a displayed column --
        # dropped before writing, same as the real page.
        _gl_summary_display = _gl_summary.drop(columns=["Mapping Required"], errors="ignore")
        _store_summary.to_excel(w, index=False, sheet_name="Summary", startrow=2, startcol=0)
        _gl_summary_display.to_excel(w, index=False, sheet_name="Summary", startrow=2, startcol=6)
        _ws = w.sheets["Summary"]
        _ws["A1"] = "STORE-WISE SUMMARY"
        _ws["G1"] = "GL-WISE SUMMARY"

        _store_total_row = 4 + len(_store_summary)
        _gl_total_row = 4 + len(_gl_summary)

        if not _store_summary.empty:
            _ws.cell(row=_store_total_row, column=1, value="TOTAL")
            _ws.cell(row=_store_total_row, column=2, value=float(_store_summary["POS Total"].sum()))
            _ws.cell(row=_store_total_row, column=3, value=float(_store_summary["GL Total"].sum()))
            _ws.cell(row=_store_total_row, column=4, value=float(_store_summary["Difference"].sum()))
        if not _gl_summary.empty:
            _ws.cell(row=_gl_total_row, column=7, value="TOTAL")
            _ws.cell(row=_gl_total_row, column=9, value=float(_gl_summary["GL Total"].sum()))
            _ws.cell(row=_gl_total_row, column=10, value=float(_gl_summary["POS Total"].sum()))
            _ws.cell(row=_gl_total_row, column=11, value=float(_gl_summary["Difference"].sum()))

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        _ok_fill = PatternFill("solid", fgColor="C6EFCE")
        _review_fill = PatternFill("solid", fgColor="FFC7CE")

        if not _store_summary.empty:
            for _r in range(4, _store_total_row):
                _cell = _ws.cell(row=_r, column=5)
                if _cell.value == "OK":
                    _cell.fill = _ok_fill
                elif _cell.value == "REVIEW":
                    _cell.fill = _review_fill
        if not _gl_summary.empty:
            for _r in range(4, _gl_total_row):
                _cell = _ws.cell(row=_r, column=12)
                if _cell.value == "OK":
                    _cell.fill = _ok_fill
                elif _cell.value == "REVIEW":
                    _cell.fill = _review_fill
    return b.getvalue(), _store_total_row, _gl_total_row


import openpyxl

data, store_total_row, gl_total_row = _write_summary_sheet(r)
wb = openpyxl.load_workbook(io.BytesIO(data))
ws = wb["Summary"]

# Conditional formatting: row 4 (store 601, exact match) must be green;
# row 6 (store 603, real gap) must be pink.
assert ws.cell(row=4, column=5).value == "OK"
assert ws.cell(row=4, column=5).fill.fgColor.rgb == "00C6EFCE"
assert ws.cell(row=6, column=5).value == "REVIEW"
assert ws.cell(row=6, column=5).fill.fgColor.rgb == "00FFC7CE"
print("[PASS] Store-wise Status cells are colored correctly: OK -> green, REVIEW -> pink.")

assert ws.cell(row=4, column=12).value == "REVIEW"
assert ws.cell(row=4, column=12).fill.fgColor.rgb == "00FFC7CE"
print("[PASS] GL-wise Status cell is colored correctly.")

# TOTAL rows: must equal the independently-verified sums.
assert ws.cell(row=store_total_row, column=1).value == "TOTAL"
assert round(float(ws.cell(row=store_total_row, column=2).value), 2) == 3500.00
assert round(float(ws.cell(row=store_total_row, column=3).value), 2) == 3000.01
assert round(float(ws.cell(row=store_total_row, column=4).value), 2) == 499.99
print("[PASS] Store-wise TOTAL row shows the correct summed POS/GL/Difference.")

assert ws.cell(row=gl_total_row, column=7).value == "TOTAL"
assert round(float(ws.cell(row=gl_total_row, column=9).value), 2) == 3000.01
assert round(float(ws.cell(row=gl_total_row, column=10).value), 2) == 3500.00
assert round(float(ws.cell(row=gl_total_row, column=11).value), 2) == 499.99
print("[PASS] GL-wise TOTAL row shows the correct summed GL/POS/Difference.")

assert round(float(ws.cell(row=store_total_row, column=4).value), 2) == round(
    float(ws.cell(row=gl_total_row, column=11).value), 2
), "Store-wise and GL-wise TOTAL Difference must match -- same internal tie-out as before"
print("[PASS] Store-wise TOTAL Difference and GL-wise TOTAL Difference agree with each other, "
      "same tie-out property already proven for the underlying tables.")

# ---------------------------------------------------------------------
# Empty case: writer must not crash and must not write a stray TOTAL row.
# ---------------------------------------------------------------------
data_empty, _, _ = _write_summary_sheet(r_empty)
wb_empty = openpyxl.load_workbook(io.BytesIO(data_empty))
ws_empty = wb_empty["Summary"]
assert ws_empty.cell(row=4, column=1).value is None
print("[PASS] Empty store_summary/gl_summary: Excel writer does not crash and writes no "
      "stray TOTAL row.")

print("REGRESSION POS GL V44 STATUS FORMATTING TOTALS PASS")
