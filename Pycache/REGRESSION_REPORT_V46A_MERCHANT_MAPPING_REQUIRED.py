# Permanent regression for V46A Merchant Mapping Required visibility.
# Run from repo root:
#   python REGRESSION_REPORT_V46A_MERCHANT_MAPPING_REQUIRED.py

from __future__ import annotations
import io
import pandas as pd
import openpyxl
import report_export as rexp

LABEL = "⚠ MERCHANT MAPPING REQUIRED"
FLAG = "Merchant Mapping Required"
ORANGE = "00FCE4D6"


def _blank_tx_row(payment: str, amount: float, auth: str, commission: float = 0.0):
    r = {c: "" for c in rexp.TX_COLUMNS}
    for p in rexp.PAYMENTS:
        r[f"D365 {p}"] = 0.0
        r[f"POS {p}"] = 0.0
        r[f"Diff {p}"] = 0.0
    r.update({
        "Store Code": "",
        "Date": pd.NaT,
        "Receipt ID": "",
        "Auth Code": auth,
        "D365 Source Type": "",
        "Cash Amount": 0.0,
        "Reservation Cash": 0.0,
        "Order Balance": 0.0,
        "Reservation Report Total": 0.0,
        "D365 Total": 0.0,
        "POS Total": amount,
        "Total Difference": -amount,
        "D365 Tender": "",
        "POS Tender": payment,
        "POS Date": pd.Timestamp("2026-07-22"),
        "Posting Date": pd.Timestamp("2026-07-23"),
        "Terminal ID": "T-UNMAPPED",
        "Commission": commission,
        "VAT": 0.0,
        "Net Amount": amount - commission,
        "Source": "synthetic_pos.xlsx",
        "Status": "Merchant Mapping Required",
        "Remarks": "Synthetic unresolved merchant/store mapping row.",
    })
    if payment in rexp.PAYMENTS:
        r[f"POS {payment}"] = amount
        r[f"Diff {payment}"] = -amount
    return r


def _mapped_tx_row(store: str, payment: str, amount: float, auth: str):
    r = {c: "" for c in rexp.TX_COLUMNS}
    for p in rexp.PAYMENTS:
        r[f"D365 {p}"] = 0.0
        r[f"POS {p}"] = 0.0
        r[f"Diff {p}"] = 0.0
    r.update({
        "Store Code": store,
        "Date": pd.Timestamp("2026-07-22"),
        "Receipt ID": "R-" + auth,
        "Auth Code": auth,
        "D365 Source Type": "Store Tender",
        "Cash Amount": 0.0,
        "Reservation Cash": 0.0,
        "Order Balance": 0.0,
        "Reservation Report Total": 0.0,
        "D365 Total": amount,
        "POS Total": amount,
        "Total Difference": 0.0,
        "D365 Tender": payment,
        "POS Tender": payment,
        "POS Date": pd.Timestamp("2026-07-22"),
        "Posting Date": pd.Timestamp("2026-07-23"),
        "Terminal ID": "T-MAPPED",
        "Commission": 0.0,
        "VAT": 0.0,
        "Net Amount": amount,
        "Source": "synthetic_pos.xlsx",
        "Status": "Matched",
        "Remarks": "Synthetic mapped row.",
    })
    if payment in rexp.PAYMENTS:
        r[f"D365 {payment}"] = amount
        r[f"POS {payment}"] = amount
        r[f"Diff {payment}"] = 0.0
    return r


tx = pd.DataFrame([
    _blank_tx_row("MADA", 228_544.39, "A1", 100.00),
    _blank_tx_row("VISA", 173_900.37, "A2", 50.00),
    _blank_tx_row("MASTER", 169_900.36, "A3", 25.00),
    _mapped_tx_row("601", "MADA", 1_000.00, "B1"),
], columns=rexp.TX_COLUMNS)

expected_unmapped = round(228_544.39 + 173_900.37 + 169_900.36, 2)

store = rexp.store_summary(tx)
assert FLAG in store.columns
assert LABEL in set(store["Store Code"].astype(str))
u = store[store["Store Code"].astype(str) == LABEL].iloc[0]
assert bool(u[FLAG]) is True
assert round(float(u["POS Total"]), 2) == expected_unmapped
assert round(float(u["D365 Total"]), 2) == 0.00
assert round(float(u["Total Difference"]), 2) == -expected_unmapped
print("[PASS] Store_Summary labels unresolved merchant activity and preserves amounts.")

m = store[store["Store Code"].astype(str) == "601"].iloc[0]
assert bool(m[FLAG]) is False
assert round(float(m["POS Total"]), 2) == 1000.00
assert round(float(m["D365 Total"]), 2) == 1000.00
print("[PASS] Normal mapped store remains unchanged.")

sett = rexp.settlement_commission(tx)
assert FLAG in sett.columns
unmapped_sett = sett[sett["Store Code"].astype(str) == LABEL].copy()
assert len(unmapped_sett) == 3
assert set(unmapped_sett["Payment Type"]) == {"MADA", "VISA", "MASTER"}
assert unmapped_sett[FLAG].map(bool).all()
assert round(float(unmapped_sett["Gross Amount"].sum()), 2) == expected_unmapped
print("[PASS] Settlement_Commission preserves payment-type split and flags all unresolved rows.")

unmatched_pos = pd.DataFrame([
    {"POS Store": "", "POS Payment": "MADA", "POS Amount": 228_544.39, "Auth Code": "A1",
     "POS Date": pd.Timestamp("2026-07-22"), "Posting Date": pd.Timestamp("2026-07-23"),
     "Terminal ID": "T-UNMAPPED", "Commission": 100.0, "VAT": 0.0, "Net Amount": 228_444.39,
     "Source File": "synthetic_pos.xlsx", "Exception Status": "Merchant Mapping Required",
     "Reason": "Synthetic unresolved merchant/store mapping row."},
    {"POS Store": "", "POS Payment": "VISA", "POS Amount": 173_900.37, "Auth Code": "A2",
     "POS Date": pd.Timestamp("2026-07-22"), "Posting Date": pd.Timestamp("2026-07-23"),
     "Terminal ID": "T-UNMAPPED", "Commission": 50.0, "VAT": 0.0, "Net Amount": 173_850.37,
     "Source File": "synthetic_pos.xlsx", "Exception Status": "Merchant Mapping Required",
     "Reason": "Synthetic unresolved merchant/store mapping row."},
    {"POS Store": "", "POS Payment": "MASTER", "POS Amount": 169_900.36, "Auth Code": "A3",
     "POS Date": pd.Timestamp("2026-07-22"), "Posting Date": pd.Timestamp("2026-07-23"),
     "Terminal ID": "T-UNMAPPED", "Commission": 25.0, "VAT": 0.0, "Net Amount": 169_875.36,
     "Source File": "synthetic_pos.xlsx", "Exception Status": "Merchant Mapping Required",
     "Reason": "Synthetic unresolved merchant/store mapping row."},
])

matched = pd.DataFrame([{
    "Store Code": "601", "Date": pd.Timestamp("2026-07-22"), "Receipt ID": "R-B1",
    "Auth Code": "B1", "Payment Type": "MADA", "D365 Amount": 1000.0, "POS Amount": 1000.0,
    "POS Date": pd.Timestamp("2026-07-22"), "Posting Date": pd.Timestamp("2026-07-23"),
    "Terminal ID": "T-MAPPED", "Commission": 0.0, "VAT": 0.0, "Net Amount": 1000.0,
    "Source File": "synthetic_pos.xlsx", "Status": "Matched",
}])

result = {
    "matched": matched,
    "cash_transactions": pd.DataFrame(),
    "unmatched_sales": pd.DataFrame(),
    "unmatched_pos": unmatched_pos,
}

xlsx = rexp.create_reconciliation_pack(result, tolerance=1.0)
wb = openpyxl.load_workbook(io.BytesIO(xlsx))
ws_store = wb["Store_Summary"]
ws_sett = wb["Settlement_Commission"]

assert ws_store.max_column == 35, f"Store_Summary width drift: {ws_store.max_column}"
assert ws_sett.max_column == 7, f"Settlement_Commission width drift: {ws_sett.max_column}"
print("[PASS] Visible widths unchanged: Store_Summary=35, Settlement_Commission=7.")

for ws in (ws_store, ws_sett):
    vals = [cell.value for row in ws.iter_rows() for cell in row]
    assert FLAG not in vals
    assert not any(isinstance(v, bool) for v in vals)
print("[PASS] Raw Merchant Mapping Required flag never leaks into visible Excel.")

def _find_rows(ws, value):
    return [r for r in range(1, ws.max_row + 1) if ws.cell(r, 1).value == value]

store_rows = _find_rows(ws_store, LABEL)
assert len(store_rows) == 1
sr = store_rows[0]
for c in range(1, ws_store.max_column + 1):
    assert ws_store.cell(sr, c).fill.fgColor.rgb == ORANGE
assert ws_store.cell(sr, 1).font.bold is True
print("[PASS] Store_Summary mapping-required row is orange across all visible columns.")

sett_rows = _find_rows(ws_sett, LABEL)
assert len(sett_rows) == 3
for rr in sett_rows:
    for c in range(1, ws_sett.max_column + 1):
        assert ws_sett.cell(rr, c).fill.fgColor.rgb == ORANGE
    assert ws_sett.cell(rr, 1).font.bold is True
print("[PASS] Settlement_Commission mapping-required rows are orange across all visible columns.")

assert "Payment_Summary" in wb.sheetnames
assert "Settlement_Delay" in wb.sheetnames
print("[PASS] Existing report sheets remain present.")

print("REGRESSION REPORT V46A MERCHANT MAPPING REQUIRED PASS")
