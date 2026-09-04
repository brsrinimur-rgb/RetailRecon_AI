"""
REGRESSION_POS_GL_V45_MAPPING_REQUIRED_FLAG.py

Item #5 from the improvement list: flag UNMAPPED rows prominently in the
GL-wise Summary instead of letting them look like just another GL account
with a variance.

  - gl_summary gets a new "Mapping Required" boolean column (True only for
    the two "no mapping resolved" sentinel cases -- an unmapped POS
    provider or an unmapped GL account -- NOT for a real, valid GL account
    that simply lacks a friendly name in core.D365_CLEARING_ACCOUNT_MAP;
    that's a different, lower-severity situation and deliberately not
    flagged the same way).
  - GL Name for those two cases is prefixed "\u26a0 MAPPING REQUIRED \u2014 ...".
  - The Excel writer highlights the WHOLE row in orange for a
    Mapping-Required row, overriding the plain OK/REVIEW coloring for that
    row -- and never writes the raw boolean column itself to the sheet.

Run: python3 REGRESSION_POS_GL_V45_MAPPING_REQUIRED_FLAG.py
"""
from pathlib import Path
import sys
import io
import pandas as pd

root = Path(__file__).resolve().parent
sys.path.insert(0, str(root))
from logic import pos_gl_reconciliation as pgl

# ---------------------------------------------------------------------
# 1. Mapping Required flag and label, correctly scoped.
# ---------------------------------------------------------------------
pos = pd.DataFrame([
    {"store_code": "601", "pos_date": "2026-07-01", "pos_amount": 300.00, "provider": "STCPAY",
     "reference": "C1", "auth_code": "C1", "source_row": 1, "merchant_id": "M1"},
])
gl = pd.DataFrame([
    {"store_code": "601", "gl_date": "2026-07-01", "gl_signed_amount": 300.00, "gl_amount": 300.00,
     "main_account": "11020950", "voucher": "V1", "journal": "J1", "source_row": 1, "source_file": "gl.xlsx"},
])
r = pgl.reconcile_pos_to_gl_by_bucket(pos, gl)
gls = r["gl_summary"].set_index("GL Account")

assert "Mapping Required" in r["gl_summary"].columns
assert bool(gls.loc["UNMAPPED", "Mapping Required"]) is True
assert "\u26a0 MAPPING REQUIRED" in gls.loc["UNMAPPED", "GL Name"]
assert "Unmapped POS Provider" in gls.loc["UNMAPPED", "GL Name"]
print("[PASS] The unmapped-provider row is flagged Mapping Required=True with the "
      "prominent GL Name label.")

# A real GL account that's simply not in the friendly-name lookup table is
# a DIFFERENT, lower-severity case -- must NOT be flagged the same way.
assert bool(gls.loc["11020950", "Mapping Required"]) is False
print("[PASS] A real GL account missing only a friendly display name is correctly "
      "NOT flagged as Mapping Required -- that's a different, lower-severity case.")

# Empty case: column must exist, just empty.
empty_pos = pd.DataFrame(columns=["store_code", "pos_date", "pos_amount", "provider",
                                   "reference", "auth_code", "source_row", "merchant_id"])
empty_gl = pd.DataFrame(columns=["store_code", "gl_date", "gl_signed_amount", "gl_amount",
                                  "main_account", "voucher", "journal", "source_row", "source_file"])
r_empty = pgl.reconcile_pos_to_gl_by_bucket(empty_pos, empty_gl)
assert "Mapping Required" in r_empty["gl_summary"].columns
assert r_empty["gl_summary"].empty
print("[PASS] Mapping Required column present even when the result is empty.")

# ---------------------------------------------------------------------
# 2. Excel writer: whole-row orange highlight for Mapping Required rows,
#    raw boolean column never written to the visible sheet.
# ---------------------------------------------------------------------
def _write_summary_sheet(r):
    b = io.BytesIO()
    with pd.ExcelWriter(b, engine="openpyxl") as w:
        _store_summary = r.get("store_summary", pd.DataFrame())
        _gl_summary = r.get("gl_summary", pd.DataFrame())
        _gl_mapping_required = (
            _gl_summary["Mapping Required"].tolist() if "Mapping Required" in _gl_summary.columns else []
        )
        _gl_summary_display = _gl_summary.drop(columns=["Mapping Required"], errors="ignore")
        _store_summary.to_excel(w, index=False, sheet_name="Summary", startrow=2, startcol=0)
        _gl_summary_display.to_excel(w, index=False, sheet_name="Summary", startrow=2, startcol=6)
        _ws = w.sheets["Summary"]

        from openpyxl.styles import Font, PatternFill

        if not _gl_summary.empty and _gl_mapping_required:
            _mapping_fill = PatternFill("solid", fgColor="FCE4D6")
            _mapping_font = Font(color="974706", bold=True)
            for _offset, _needs_mapping in enumerate(_gl_mapping_required):
                if not _needs_mapping:
                    continue
                _r = 4 + _offset
                for _c in range(7, 13):
                    _cell = _ws.cell(row=_r, column=_c)
                    _cell.fill = _mapping_fill
                    if _c in (7, 8):
                        _cell.font = _mapping_font
    return b.getvalue()


import openpyxl

data = _write_summary_sheet(r)
wb = openpyxl.load_workbook(io.BytesIO(data))
ws = wb["Summary"]

# GL rows are alphabetically sorted by key: "11020950" sorts before
# "UNMAPPED" (digit < letter), so row 4 = 11020950, row 5 = UNMAPPED.
assert ws.cell(row=4, column=1 + 6).value == "11020950"
assert ws.cell(row=4, column=1 + 6).fill.fgColor.rgb == "00000000", "non-flagged row must not be highlighted"
assert ws.cell(row=5, column=1 + 6).value == "UNMAPPED"
assert ws.cell(row=5, column=1 + 6).fill.fgColor.rgb == "00FCE4D6"
assert ws.cell(row=5, column=3 + 6).fill.fgColor.rgb == "00FCE4D6", "highlight must span the whole row, not just GL Account"
assert ws.cell(row=5, column=1 + 6).font.bold is True
print("[PASS] Excel export: the Mapping-Required row is highlighted orange across the "
      "whole row; the unrelated row above it is untouched.")

# The raw boolean column must never appear as visible data on the sheet.
found_boolean_cell = False
for row in ws.iter_rows():
    for cell in row:
        if isinstance(cell.value, bool):
            found_boolean_cell = True
assert not found_boolean_cell, "the raw Mapping Required boolean must never be written to the visible sheet"
print("[PASS] The raw 'Mapping Required' boolean column is never written to the visible sheet.")

print("REGRESSION POS GL V45 MAPPING REQUIRED FLAG PASS")
