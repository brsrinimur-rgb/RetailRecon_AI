
import io, zipfile, hashlib, uuid
from pathlib import Path
import pandas as pd
import streamlit as st
import auth, theme
from logic.pos_gl_reconciliation import (
    build_pos_dataset, build_gl_dataset,
    reconcile_pos_to_gl, reconcile_pos_to_gl_by_bucket,
    detect_exact_duplicate_files,
    detect_probable_duplicate_pos, detect_probable_duplicate_gl,
    detect_content_duplicate_pos, detect_content_duplicate_gl,
    validate_pos_completeness, validate_gl_completeness, upload_control_summary,
    detect_incomplete_pos_provider_coverage, validate_gl_sign_convention,
)
from logic import swap_tracking, provider_gl_mapping
from datetime import datetime, timezone
import core

def _sig(pairs):
    """(filename, sha256(bytes)) identity for a set of uploaded files --
    V40 item 4: filename alone can't tell a same-named re-upload with
    different content apart from the file already validated."""
    return [(n, hashlib.sha256(d).hexdigest()) for n, d in pairs]

def _deterministic_run_signature(pos_sig, gl_sig):
    """V42 item 1: replaces Python's built-in hash(). hash() is salted per
    process (PYTHONHASHSEED) -- the exact same upload set produces a
    DIFFERENT Run Signature after every app restart, which defeats the
    entire point of a signature meant to prove two runs used identical
    inputs. This is a plain SHA256 over a sorted, canonical serialization
    of both signatures, so it's stable across restarts, processes, and
    machines -- the same files always produce the same Run Signature."""
    canonical = repr((sorted(pos_sig), sorted(gl_sig)))
    return hashlib.sha256(canonical.encode("utf-8")).hexdigest()

def _build_import_audit(pos_audit, gl_audit):
    """V42 items 2+3: shared by both the Validate display block and the RUN
    handler so 'what got audited' is computed the same way in both places,
    from whichever pos_audit/gl_audit are in hand at that moment -- rather
    than the RUN handler silently relying on session_state left over from a
    Validate click on a possibly different set of files (see the RUN
    handler's own comment for why that was a real gap)."""
    parts = [a.assign(**{"Upload": u}) for a, u in [(pos_audit, "POS"), (gl_audit, "GL")] if a is not None and not a.empty]
    return pd.concat(parts, ignore_index=True) if parts else pd.DataFrame()

st.set_page_config(page_title="POS → D365 GL Reconciliation", layout="wide", page_icon="🧾")
auth.require_login({"Admin","Finance Manager","Finance Checker"})
auth.render_user_sidebar()
st.markdown(theme.global_css(), unsafe_allow_html=True)
st.markdown(theme.top_banner("RETAIL CONTROL TOWER","POS Statement → D365 GL Reconciliation"), unsafe_allow_html=True)

st.title("🧾 POS Statement → D365 GL Reconciliation")
st.caption("Daily bulk reconciliation: multiple POS files + multiple D365 GL account dumps. Store Tender is NOT used.")
st.info(
    "Accounting authority: POS Statement Amount ↔ D365 GL Amount, summed per Store + Date. "
    "Store Code and Date identify the bucket; Amount alone proves or rejects the match. "
    "Merchant ID, Provider, Terminal and Auth Code stay available as investigation detail, "
    "never as part of the matching key."
)

def read_one_bytes(name, data):
    lname=name.lower()
    if lname.endswith(".csv"):
        return pd.read_csv(io.BytesIO(data))
    return pd.concat(
        [pd.read_excel(io.BytesIO(data), sheet_name=s) for s in pd.ExcelFile(io.BytesIO(data)).sheet_names],
        ignore_index=True
    )

def expand_zip(zf):
    result=[]
    with zipfile.ZipFile(io.BytesIO(zf.getvalue())) as z:
        for info in z.infolist():
            if info.is_dir(): continue
            name=info.filename
            if name.lower().endswith((".xlsx",".xls",".csv")):
                result.append((Path(name).name,z.read(info)))
    return result

st.subheader("1. POS Statements — BULK EXCEL UPLOAD")
st.caption("Upload many daily POS Excel/CSV files together. There is no one-file limit.")

pos_uploads=st.file_uploader(
    "📤 UPLOAD MULTIPLE POS EXCEL FILES",
    type=["xlsx","xls","csv"],
    accept_multiple_files=True,
    key="v54_pos_multi",
    help="In the Windows picker, hold Ctrl or Shift to select several files. You can also drag multiple files into this box."
)
pos_zip=st.file_uploader(
    "OR upload ONE ZIP containing all POS files",
    type=["zip"],
    accept_multiple_files=False,
    key="v54_pos_zip"
)

pos_pairs=[]
for f in (pos_uploads or []):
    pos_pairs.append((f.name,f.getvalue()))
if pos_zip:
    pos_pairs.extend(expand_zip(pos_zip))

# V40 item 3: every uploaded file is kept, including two files sharing a
# filename -- a filename-only pre-filter used to silently drop a same-named
# duplicate BEFORE SHA256/content duplicate detection below ever saw it.
# None of the three duplicate tiers auto-excludes a file; they're warnings
# for a human to resolve, so keeping every upload here only affects what
# gets FLAGGED, not what gets summed.

if pos_pairs:
    st.success(f"POS files loaded: {len(pos_pairs)}")
    with st.expander("View POS files", expanded=False):
        st.write("\n".join(f"• {n}" for n,_ in pos_pairs))
else:
    st.info("POS: upload multiple Excel files above, or upload one ZIP batch.")

st.subheader("2. D365 GL — BULK EXCEL UPLOAD")
st.caption("Upload all D365 GL account dumps together. 8, 20, 50+ GL accounts are supported.")

gl_uploads=st.file_uploader(
    "📤 UPLOAD MULTIPLE D365 GL EXCEL FILES",
    type=["xlsx","xls","csv"],
    accept_multiple_files=True,
    key="v54_gl_multi",
    help="In the Windows picker, hold Ctrl or Shift to select several files. You can also drag multiple files into this box."
)
gl_zip=st.file_uploader(
    "OR upload ONE ZIP containing all D365 GL files",
    type=["zip"],
    accept_multiple_files=False,
    key="v54_gl_zip"
)

gl_pairs=[]
for f in (gl_uploads or []):
    gl_pairs.append((f.name,f.getvalue()))
if gl_zip:
    gl_pairs.extend(expand_zip(gl_zip))

# V40 item 3: same reasoning as the POS block above -- keep every upload.

if gl_pairs:
    st.success(f"D365 GL files loaded: {len(gl_pairs)}")
    with st.expander("View D365 GL files", expanded=False):
        st.write("\n".join(f"• {n}" for n,_ in gl_pairs))
else:
    st.info("D365 GL: upload multiple Excel files above, or upload one ZIP batch.")

# V41 item 1: compute the CURRENT upload identity once, up front, and reuse
# it everywhere below (Validate auto-invalidation, the RUN button's
# same_upload check, the stale-result banner) instead of recomputing
# _sig(pos_pairs)/_sig(gl_pairs) -- itself a SHA256 hash of every file --
# repeatedly.
current_pos_sig = _sig(pos_pairs)
current_gl_sig = _sig(gl_pairs)

st.subheader("3. Validate")
st.caption("Checks uploads for duplicate files, missing Store Code/Date, and shows what's actually loaded -- before you commit to a full reconciliation run.")

# V41 item 1: auto-invalidate a stale Validate result. Before this fix,
# v53_validated stayed True after the user changed an upload (added/removed
# a file, or swapped one for a different one under the same name), so the
# Validate summary below could keep showing normalized data built from
# files that are no longer what's in the uploader -- with nothing on
# screen saying so.
if st.session_state.get("v53_validated") and (
    st.session_state.get("v53_pos_pairs_sig") != current_pos_sig or
    st.session_state.get("v53_gl_pairs_sig") != current_gl_sig
):
    st.session_state.v53_validated = False
    st.warning("⚠️ UPLOADS CHANGED — VALIDATE AND RUN AGAIN. The Validate summary below was cleared because the POS/GL files currently loaded no longer match what you last validated.")

if st.button("🔍 VALIDATE UPLOADS", use_container_width=True):
    if not pos_pairs or not gl_pairs:
        st.error("Load at least one POS file and one GL file (or ZIP batch) before validating.")
        st.stop()
    with st.spinner("Reading and normalizing uploads..."):
        npos, pos_audit = build_pos_dataset(pos_pairs)
        ngl, gl_audit = build_gl_dataset(gl_pairs)
        st.session_state.v53_pos_ds = npos
        st.session_state.v53_gl_ds = ngl
        st.session_state.v53_pos_audit = pos_audit
        st.session_state.v53_gl_audit = gl_audit
        # V40 item 4: identity is (filename, sha256(bytes)), not filename
        # alone -- a same-named re-upload with different content used to be
        # silently treated as "the same upload" and could reuse stale
        # cached normalized data.
        st.session_state.v53_pos_pairs_sig = current_pos_sig
        st.session_state.v53_gl_pairs_sig = current_gl_sig
        st.session_state.v53_validated = True

if st.session_state.get("v53_validated"):
    npos = st.session_state.v53_pos_ds
    ngl = st.session_state.v53_gl_ds
    pos_audit = st.session_state.get("v53_pos_audit")
    gl_audit = st.session_state.get("v53_gl_audit")

    exact_pos = detect_exact_duplicate_files(pos_pairs)
    exact_gl = detect_exact_duplicate_files(gl_pairs)
    probable_pos = detect_probable_duplicate_pos(npos)
    probable_gl = detect_probable_duplicate_gl(ngl)
    possible_pos = detect_content_duplicate_pos(npos)
    possible_gl = detect_content_duplicate_gl(ngl)
    all_exact = exact_pos + exact_gl
    all_probable = probable_pos + probable_gl
    all_possible = possible_pos + possible_gl

    # V40 item 7: severity-differentiated instead of one undifferentiated
    # red banner for all three tiers -- Possible is the weakest evidence
    # (same row count/total/date range only, which two different legitimate
    # files could share by coincidence) and should never be worded as if
    # double-counting definitely occurred.
    if all_exact:
        st.error(
            f"🔴 {len(all_exact)} EXACT duplicate file(s) detected -- byte-identical content, almost "
            f"certainly the same file uploaded twice. Totals will double-count until this is resolved. "
            f"No file is auto-excluded; confirm and remove the duplicate before running."
        )
        for d in all_exact:
            st.write(f"• `{d['file_a']}` and `{d['file_b']}` -- {d['reason']}")
    if all_probable:
        st.warning(
            f"🟠 {len(all_probable)} PROBABLE duplicate file(s) detected -- at least 90% of individual "
            f"transactions (Store+Date+Amount) match between two files. Very likely the same data "
            f"uploaded twice under different filenames; double-check before running."
        )
        for d in all_probable:
            st.write(f"• `{d['file_a']}` and `{d['file_b']}` -- {d['reason']}")
    if all_possible:
        st.info(
            f"🟡 {len(all_possible)} POSSIBLE duplicate file(s) noted -- same row count, total amount "
            f"and date range only. This can happen between two different, legitimate files by "
            f"coincidence; it is NOT evidence that double-counting has actually occurred, only a "
            f"prompt to take a look."
        )
        for d in all_possible:
            st.write(f"• `{d['file_a']}` and `{d['file_b']}` -- {d['reason']}")
    if not (all_exact or all_probable or all_possible):
        st.success("✅ No duplicate files detected among the uploads.")

    coverage_warnings = detect_incomplete_pos_provider_coverage(npos, ngl)
    if coverage_warnings:
        st.warning(f"⚠️ UPLOAD COMPLETENESS: {len(coverage_warnings)} Store/provider combination(s) posted in D365 GL have no matching POS file uploaded. This is a separate control from accounting exceptions below -- it means the upload is likely incomplete, not that GL posted something wrong.")
        with st.expander("View missing POS provider coverage", expanded=True):
            for w in coverage_warnings:
                st.write(f"• {w['message']}")
    else:
        st.success("✅ Upload completeness: every GL clearing account has a matching POS provider file for that store.")

    sign_check = validate_gl_sign_convention(npos, ngl)
    if sign_check.get("checked"):
        if sign_check["suspected_inverted"]:
            st.warning(f"⚠️ GL SIGN CONVENTION: {sign_check['message']}")
        else:
            st.caption(f"✅ Sign check: {sign_check['message']}")

    comp = validate_pos_completeness(npos)
    if comp["missing_store"] or comp["missing_date"] or comp["missing_amount"]:
        st.warning(
            f"POS completeness: {comp['clean_rows']:,}/{comp['total_rows']:,} rows usable. "
            f"{comp['missing_store']:,} missing Store Code, {comp['missing_date']:,} missing Date, "
            f"{comp['missing_amount']:,} missing Amount -- these rows can't be bucketed and will show as "
            f"Identifier Mismatch / POS Data Incomplete."
        )
        with st.expander("View rows with missing Store Code / Date / Amount", expanded=False):
            st.dataframe(comp["sample"], use_container_width=True, hide_index=True)
    else:
        st.success(f"✅ POS completeness: all {comp['total_rows']:,} rows have a Store Code, Date and Amount.")

    # V40 item 9: GL-side equivalent of the POS completeness check above --
    # GL never had a dedicated preflight for this before.
    gl_comp = validate_gl_completeness(ngl)
    if gl_comp["missing_store"] or gl_comp["missing_date"] or gl_comp["missing_account"] or gl_comp["missing_amount"]:
        st.warning(
            f"GL completeness: {gl_comp['clean_rows']:,}/{gl_comp['total_rows']:,} rows usable. "
            f"{gl_comp['missing_store']:,} missing Store Code, {gl_comp['missing_date']:,} missing GL Date, "
            f"{gl_comp['missing_account']:,} missing Main Account, {gl_comp['missing_amount']:,} missing "
            f"Signed Amount -- these rows can't be bucketed and are excluded from reconciliation totals "
            f"and reported below, not silently dropped."
        )
        with st.expander("View GL rows with missing Store Code / Date / Main Account / Signed Amount", expanded=False):
            st.dataframe(gl_comp["sample"], use_container_width=True, hide_index=True)
    else:
        st.success(f"✅ GL completeness: all {gl_comp['total_rows']:,} rows have a Store Code, GL Date, Main Account and Signed Amount.")

    ctrl = upload_control_summary(pos_pairs, gl_pairs, npos, ngl)
    st.subheader("Upload Control Summary")
    a,b,c,d = st.columns(4)
    a.metric("POS Files", ctrl["pos_files"]); b.metric("GL Files", ctrl["gl_files"])
    c.metric("POS Rows", f"{ctrl['pos_rows']:,}"); d.metric("GL Rows", f"{ctrl['gl_rows']:,}")
    a,b,c = st.columns(3)
    a.metric("POS Stores", len(ctrl["pos_stores"])); b.metric("GL Stores", len(ctrl["gl_stores"])); c.metric("GL Clearing Accounts", len(ctrl["gl_accounts"]))

    period = ctrl["period"]
    st.write(f"**POS date range:** {period['pos_date_min']} → {period['pos_date_max']} ({period['pos_distinct_dates']} distinct dates)")
    st.write(f"**GL date range:** {period['gl_date_min']} → {period['gl_date_max']} ({period['gl_distinct_dates']} distinct dates)")
    if period["dates_pos_only"]:
        st.warning(f"{len(period['dates_pos_only'])} date(s) have POS activity but no GL file covering them: " + ", ".join(str(d.date()) for d in period["dates_pos_only"][:15]) + (" ..." if len(period["dates_pos_only"])>15 else ""))
    if period["dates_gl_only"]:
        st.warning(f"{len(period['dates_gl_only'])} date(s) have GL activity but no POS file covering them: " + ", ".join(str(d.date()) for d in period["dates_gl_only"][:15]) + (" ..." if len(period["dates_gl_only"])>15 else ""))
    if ctrl["stores_pos_only"]:
        st.caption(f"Stores in POS only (no GL clearing activity at all): {', '.join(ctrl['stores_pos_only'])}")
    if ctrl["stores_gl_only"]:
        st.caption(f"Stores in GL only (no POS activity at all): {', '.join(ctrl['stores_gl_only'])}")

    # V41 item 4: Import Audit / Quarantine -- one row per (file, sheet)
    # actually attempted while building npos/ngl, whether it succeeded, was
    # intentionally skipped (settlement/payout sheet), or was quarantined
    # (couldn't be parsed at all). Before this pass, build_pos_dataset()/
    # build_gl_dataset() had a broad `except Exception: continue` -- a file
    # or sheet that failed to parse simply vanished from the run with no
    # visible trace anywhere in the app.
    import_audit = _build_import_audit(pos_audit, gl_audit)
    # V42 items 2+3: this session_state copy is now display-only -- a
    # convenience for showing "what did Validate last see" if this block
    # re-renders. It is NOT what the Excel export reads any more (see the
    # RUN handler and export section below): the export now always reads
    # the audit frozen onto the RESULT itself at RUN time, so a later
    # Validate click on a different set of files can never make an export
    # describe files other than the ones that actually produced it.
    st.session_state.v53_import_audit = import_audit
    st.subheader("Import Audit")
    quarantined = import_audit[import_audit["Status"] == "QUARANTINED"] if not import_audit.empty else import_audit
    if not import_audit.empty and not quarantined.empty:
        st.error(
            f"🔴 {len(quarantined)} file/sheet(s) could NOT be parsed and were quarantined -- their rows are "
            f"NOT included anywhere in this run's totals. Review the reasons below; a quarantined sheet is a "
            f"silent gap in coverage until it's fixed and re-uploaded."
        )
    else:
        st.success("✅ Every uploaded file/sheet was either loaded successfully or intentionally skipped (settlement/payout sheets) -- nothing was silently dropped.")
    with st.expander(f"View Import Audit ({len(import_audit)} file/sheet entries)", expanded=not quarantined.empty if not import_audit.empty else False):
        st.caption("Status: OK = loaded into the run. SKIPPED = a bank-side settlement/payout sheet, correctly excluded from POS matching. FILTERED = GL rows on a non-clearing account (Sales/COGS/Tax/...), correctly out of scope. QUARANTINED = could not be parsed at all -- not in this run's totals.")
        st.dataframe(import_audit, use_container_width=True, hide_index=True)

    # V42 item 6: Admin-maintained Provider -> GL Mapping master. A
    # PROVIDER MAPPING REQUIRED bucket (or a store-wide coverage warning
    # above) can now be resolved here, without a code change or redeploy --
    # this override is checked FIRST by every provider->GL resolution in
    # this app (reconcile_pos_to_gl_by_bucket's coverage split,
    # detect_incomplete_pos_provider_coverage's store-wide warning), falling
    # back to core._gl_expected_account_for_tender() unchanged when no
    # override exists for a given (provider, store).
    with st.expander("Admin: Provider → GL Mapping", expanded=False):
        st.caption(
            "Overrides checked before the built-in mapping table. Leave Store Code blank for a mapping "
            "that applies to every store (a store-specific mapping added separately still wins for that "
            "one store)."
        )
        existing = provider_gl_mapping.list_provider_gl_mappings()
        if existing:
            st.dataframe(
                pd.DataFrame(existing)[["provider", "store_code", "gl_accounts", "added_by", "added_at", "note"]]
                  .rename(columns={"provider": "Provider", "store_code": "Store Code (blank = all)",
                                    "gl_accounts": "GL Accounts", "added_by": "Added By",
                                    "added_at": "Added At (UTC)", "note": "Note"}),
                use_container_width=True, hide_index=True,
            )
        else:
            st.caption("No overrides added yet -- every provider currently resolves via core.py's built-in table.")

        acct_options = sorted(core.D365_CLEARING_ACCOUNT_MAP.keys())
        acct_labels = {a: f"{a} -- {core.D365_CLEARING_ACCOUNT_MAP[a].get('account_name', a)}" for a in acct_options}
        with st.form("v53_add_provider_mapping", clear_on_submit=True):
            fc1, fc2 = st.columns(2)
            with fc1:
                new_provider = st.text_input("Provider value (as it appears in the POS file)")
            with fc2:
                new_store = st.text_input("Store Code (optional -- blank = all stores)")
            new_accounts = st.multiselect(
                "D365 GL clearing account(s)", options=acct_options, format_func=lambda a: acct_labels.get(a, a)
            )
            new_note = st.text_input("Note (optional)")
            if st.form_submit_button("Add / Update mapping"):
                if not new_provider or not new_accounts:
                    st.error("Provider value and at least one GL account are required.")
                else:
                    ok = provider_gl_mapping.save_provider_gl_mapping(
                        new_provider, new_accounts, store_code=new_store,
                        added_by=st.session_state.get("user", {}).get("username", ""),
                        note=new_note,
                    )
                    if ok:
                        st.success(f"Saved mapping for {new_provider!r}.")
                        st.rerun()
                    else:
                        st.error("Could not save this mapping -- check the values above.")

        if existing:
            del_options = {f"{m['provider']} / {m['store_code'] or '(all stores)'}": m for m in existing}
            del_choice = st.selectbox("Remove a mapping", options=["(select)"] + list(del_options.keys()))
            if del_choice != "(select)" and st.button("Delete selected mapping"):
                m = del_options[del_choice]
                if provider_gl_mapping.delete_provider_gl_mapping(m["provider"], m["store_code"]):
                    st.success(f"Deleted mapping for {del_choice}.")
                    st.rerun()
                else:
                    st.error("Could not delete this mapping.")

st.divider()
tc1, tc2 = st.columns(2)
with tc1:
    tolerance=st.number_input("Matching tolerance per line item (SAR)",0.0,10.0,0.50,0.01,
        help="Applied per transaction/line in a bucket -- effective bucket tolerance scales with how many POS/GL rows fall into it, so summing many independently-rounded transactions doesn't manufacture false exceptions.")
with tc2:
    max_bucket_tolerance=st.number_input("Maximum bucket tolerance ceiling (SAR)",1.0,10000.0,25.00,1.0,
        help="Hard cap on the effective per-bucket tolerance. Without this, a bucket with thousands of line items could accept a real difference of thousands of SAR as MATCHED. Never allow this to grow unbounded.")

granularity=st.radio(
    "Matching granularity",
    ["Store + Date bucket (recommended)","Row-to-row (1:1) -- diagnostic / legacy"],
    index=0,
    help=(
        "Bucket (recommended): compares total POS Amount vs total D365 GL Amount per Store + Date. "
        "Store Code and Date identify the bucket; the amount alone decides match vs exception. "
        "Row-to-row (diagnostic/legacy): requires a single POS transaction to match a single GL line "
        "exactly -- only produces matches if your D365 export truly posts one GL line per POS "
        "transaction. It also does NOT have the V40/V41 control stack: no provider coverage split, "
        "no UPLOAD INCOMPLETE / PROVIDER MAPPING REQUIRED detection, no chronic-store detection, no "
        "Import Audit or run metadata, and a weaker Overall Status. Use it only to diagnose a specific "
        "row-level question, not as the primary reconciliation view."
    ),
)
settlement_lag_days=0
exclude_dup_pos=True
exclude_dup_gl=True
if granularity.startswith("Store + Date"):
    c1,c2 = st.columns(2)
    with c1:
        settlement_lag_days=st.number_input(
            "Settlement lag (days) -- shifts POS Date forward before matching to GL Date",
            0,10,0,1,
            help="Default 0: D365 GL/journal postings are normally booked the same day as the sale.",
        )
    with c2:
        st.write("Row-level duplicate exclusion")
        exclude_dup_pos=st.checkbox("Exclude duplicate POS rows from bucket sums (Store+Date+Reference/Auth+Amount)",value=True)
        # V41 item 8: label now states the full 6-field key (V40 added
        # Signed Amount) and the help text says plainly what the residual
        # risk is -- this control was already configurable, but the risk it
        # carries wasn't visible anywhere near it.
        exclude_dup_gl=st.checkbox(
            "Exclude duplicate GL rows from bucket sums (Voucher+Journal+Account+Store+Date+Amount)",
            value=True,
            help="Auto-excludes GL rows that share all six fields, keeping one copy. Two legitimate D365 lines could in theory still share all six by coincidence -- file-level duplicate detection in Validate (exact/probable/possible) is the stronger, primary control. Turn this off to compare results with GL row exclusion disabled if a run's \"Duplicate GL Excluded\" count looks high.",
        )

if st.button("RUN POS → GL RECONCILIATION",type="primary",use_container_width=True):
    if not pos_pairs or not gl_pairs:
        st.error("Load at least one POS file and one GL file (or ZIP batch) before running.")
        st.stop()

    # Each file/sheet is read and normalized on its own (real header-row
    # detection + provider-specific parsing + D365 Ledger Account dimension
    # parsing all happen per file here) instead of being concatenated into
    # one raw blob first -- that concatenation was what caused every POS
    # and GL identity field to come out blank. Reuse the Validate step's
    # already-normalized datasets when available so files aren't re-parsed.
    same_upload = (st.session_state.get("v53_pos_pairs_sig")==current_pos_sig
                   and st.session_state.get("v53_gl_pairs_sig")==current_gl_sig)
    if st.session_state.get("v53_validated") and same_upload:
        npos = st.session_state.v53_pos_ds
        ngl = st.session_state.v53_gl_ds
        pos_audit = st.session_state.get("v53_pos_audit")
        gl_audit = st.session_state.get("v53_gl_audit")
    else:
        npos, pos_audit = build_pos_dataset(pos_pairs)
        ngl, gl_audit = build_gl_dataset(gl_pairs)
        st.session_state.v53_pos_audit = pos_audit
        st.session_state.v53_gl_audit = gl_audit

    # V42 items 2+3: build Import Audit from THIS run's own pos_audit/
    # gl_audit (whichever branch above supplied them -- reused from Validate,
    # or freshly rebuilt) and freeze it onto the result below, rather than
    # relying on st.session_state.v53_import_audit. Before this fix, that
    # session_state value was set ONLY inside the Validate display block --
    # neither branch of this RUN handler ever touched it -- so the Excel
    # export could silently ship an Import Audit sheet describing a DIFFERENT
    # set of files than the ones this specific run actually reconciled
    # (e.g. Validate clicked once, files changed, RUN clicked without
    # re-validating in the "reparse" branch above).
    run_import_audit = _build_import_audit(pos_audit, gl_audit)

    # V40 item 2: a fresh run_id is generated ONLY here, inside the actual
    # button-click execution -- this block does not run on a Streamlit
    # rerun triggered by an unrelated widget, so the run_id stays stable
    # across those reruns (read back from session_state below) and only
    # changes when the user genuinely runs reconciliation again.
    run_id = str(uuid.uuid4())
    st.session_state.v53_run_id = run_id
    # V41 item 1 & 2: freeze exactly what this RUN was built from -- the
    # upload signature (to detect a later stale-result situation) and run
    # metadata (file counts, timestamp) baked into the result itself, so the
    # dashboard never has to re-derive "what was reconciled" from whatever
    # is CURRENTLY sitting in the uploader widgets.
    st.session_state.v53_result_pos_sig = current_pos_sig
    st.session_state.v53_result_gl_sig = current_gl_sig
    # V42 item 1: deterministic SHA256 signature -- Python's built-in hash()
    # is process-salted (PYTHONHASHSEED) and is NOT guaranteed stable across
    # app restarts, so the exact same upload set could previously produce a
    # different Run Signature after a redeploy, defeating its purpose as a
    # "these two runs used identical inputs" proof.
    run_signature = _deterministic_run_signature(current_pos_sig, current_gl_sig)
    run_timestamp = datetime.now(timezone.utc).isoformat(timespec="seconds")

    if granularity.startswith("Store + Date"):
        st.session_state.v53_pos_gl=reconcile_pos_to_gl_by_bucket(
            npos,ngl,tolerance,settlement_lag_days,exclude_dup_pos,exclude_dup_gl,max_bucket_tolerance,
            run_id=run_id, run_signature=run_signature, pos_file_count=len(pos_pairs),
            gl_file_count=len(gl_pairs), run_timestamp=run_timestamp,
        )
    else:
        st.session_state.v53_pos_gl=reconcile_pos_to_gl(npos,ngl,tolerance)
    # V42 items 2+3: attach directly to the result dict so the Excel export
    # (and anything else downstream) reads the audit that matches THIS
    # result, regardless of what Validate is showing by the time of export.
    st.session_state.v53_pos_gl["import_audit"] = run_import_audit

r=st.session_state.get("v53_pos_gl")
if r:
    s=r["summary"].iloc[0]
    overall=s.get("Overall Status","EXCEPTIONS REQUIRE REVIEW")
    is_bucket = "bucket_summary" in r

    # V41 item 1: stale-result banner -- the uploader can be changed after a
    # RUN without anyone clicking RUN again; before this fix the dashboard
    # below just kept showing the old result with no indication the current
    # uploads no longer match it.
    result_pos_sig = st.session_state.get("v53_result_pos_sig")
    result_gl_sig = st.session_state.get("v53_result_gl_sig")
    if result_pos_sig is not None and (result_pos_sig != current_pos_sig or result_gl_sig != current_gl_sig):
        st.error("🔴 UPLOADS CHANGED SINCE THIS RESULT WAS PRODUCED — the dashboard below reflects an earlier set of files, not what's currently loaded above. Click RUN again (after Validate, if needed) before trusting these numbers.")

    def _int_field(row, key, default=0):
        v = row.get(key, default)
        try:
            return int(v)
        except (TypeError, ValueError):
            return default

    if is_bucket:
        st.subheader("POS → D365 GL CONTROL")
        st.caption("Matched / Exceptions below are counted at the Store+Date BUCKET level -- one bucket, however many POS transactions it contains, counts once. POS/GL row counts are shown separately so the two are never confused.")
        # V41 item 2: Files Loaded reads from the run's OWN frozen metadata
        # (s["POS Files"]/s["GL Files"], set at RUN time) instead of the
        # live pos_pairs/gl_pairs lists -- those can have changed since this
        # result was produced (see the stale-result banner above), and
        # recomputing "Files Loaded" from them would then describe a
        # different population than the financial numbers next to it.
        files_loaded = _int_field(s, "POS Files") + _int_field(s, "GL Files")
        a,b,c=st.columns(3)
        a.metric("Files Loaded",f"{files_loaded}"); b.metric("POS Transaction Rows",f"{int(s['POS Transaction Rows']):,}"); c.metric("GL Line Rows",f"{int(s['GL Line Rows']):,}")
        a,b,c=st.columns(3)
        a.metric("Store-Date Buckets",f"{int(s['Store-Date Buckets']):,}"); b.metric("Matched Buckets",f"{int(s['Matched Buckets']):,}"); c.metric("Exception Buckets",f"{int(s['Exception Buckets']):,}")
        a,b,c=st.columns(3)
        a.metric("GL Not Posted (buckets)",f"{int(s['GL Not Posted']):,}"); b.metric("Duplicate POS Excluded",f"{int(s['Duplicate POS Rows Excluded']):,}"); c.metric("Duplicate GL Excluded",f"{int(s['Duplicate GL Rows Excluded']):,}")
        # V41 item 8: GL row-level duplicate auto-exclusion is a live risk
        # (two legitimate D365 lines could in theory share Voucher+Journal+
        # Account+Store+Date+Amount) -- make it visible whenever it actually
        # excluded something, not just a number sitting in a metric tile.
        if _int_field(s, "Duplicate GL Rows Excluded") > 0:
            st.caption(f"⚠️ {int(s['Duplicate GL Rows Excluded']):,} GL row(s) were auto-excluded as duplicates (Voucher+Journal+Account+Store+Date+Amount). File-level duplicate detection in Validate is the primary control -- if this count looks high relative to the files you uploaded, review the \"Store Date Buckets\" tab or turn off \"Exclude duplicate GL rows\" above and re-run to compare.")
        a,b,c=st.columns(3)
        a.metric("POS Total (SAR)",f"{s['POS Total (SAR)']:,.2f}"); b.metric("GL Total (SAR)",f"{s['GL Total (SAR)']:,.2f}"); c.metric("Net Difference (SAR)",f"{s['Net Difference (SAR)']:,.2f}")
        a,b,c=st.columns(3)
        a.metric("Extreme Variance Buckets",f"{int(s.get('Extreme Variance Buckets',0)):,}")
        b.metric("Upload Incomplete Buckets",f"{_int_field(s,'Upload Incomplete Buckets') + _int_field(s,'Provider Mapping Required Buckets')}",
                 help="Buckets where GL activity in one or more clearing accounts has no matching POS/provider file for this exact Store+Date -- once that's set aside, POS and the remaining GL agree within tolerance. Includes both UPLOAD INCOMPLETE (no such provider was uploaded at all) and PROVIDER MAPPING REQUIRED (a provider WAS uploaded but this app can't map it to a GL account) -- see the Upload Incomplete tab to tell them apart.")
        c.metric("Uncovered GL Activity (SAR)",f"{s.get('Uncovered GL Activity (SAR)',0):,.2f}",
                 help="Signed sum across every bucket -- positive and negative uncovered amounts can offset. See \"Uncovered GL Absolute Exposure\" below for the risk-sized figure.")
        a,_,_=st.columns(3)
        a.metric("Uncovered GL Absolute Exposure (SAR)",f"{s.get('Uncovered GL Absolute Exposure (SAR)',0):,.2f}",
                 help="V41: sum of ABSOLUTE uncovered GL activity across every bucket -- unlike the signed figure above, this can't be understated by positive/negative buckets cancelling out. The better KPI for how much GL activity currently has no POS/provider evidence.")
        st.caption(f"Overall: {overall} · Max bucket tolerance ceiling: SAR {s['Max Bucket Tolerance SAR']:,.2f}")
        run_id_disp = s.get("Run ID") or "(not recorded)"
        run_ts_disp = s.get("Run Timestamp") or "(not recorded)"
        st.caption(f"Run ID: `{run_id_disp}` · Run Timestamp: {run_ts_disp} (UTC) -- frozen at RUN time, V41 item 2.")

        chronic = r.get("chronic_stores")
        if chronic is not None and not chronic.empty:
            st.error(
                f"🔴 {len(chronic)} store(s) show a CHRONIC pattern -- failing on nearly every date in "
                f"this run, not just an isolated exception. This is a persistent control failure, not a "
                f"one-off data error -- review the Failure Pattern column below to see which cause "
                f"dominates for each store (a missing upload, an unmapped provider, GL postings, or a "
                f"genuine accounting variance can all look chronic, and the fix is different for each)."
            )
            st.dataframe(chronic,use_container_width=True,hide_index=True)

        dup_dates = r.get("duplicate_dates")
        if dup_dates is not None and not dup_dates.empty:
            st.error(
                f"🔴 {len(dup_dates)} date(s) show a possible SYSTEM-WIDE duplicate upload -- many "
                f"different stores excluding duplicate POS rows on the same date usually means that "
                f"day's POS export was uploaded twice."
            )
            st.dataframe(dup_dates,use_container_width=True,hide_index=True)

        # V40 item 2: run_id stays fixed across Streamlit reruns of the same
        # displayed result (set once inside the RUN button's click handler
        # above), so this call -- which does re-execute on every rerun,
        # since it has to for the cached result to stay visible -- is now
        # idempotent instead of logging a fresh sighting every rerun.
        # V42.1 compatibility fix:
        # Some deployed repos still have the pre-V40 swap_tracking.py whose
        # record_and_annotate_swaps() does not accept run_id=. Calling it
        # with the keyword raises TypeError and breaks the whole page after
        # reconciliation has already completed.
        #
        # Cache one result per actual reconciliation run so even the legacy
        # function is not called again on ordinary Streamlit reruns (which
        # would otherwise inflate Times Seen).
        _swap_run_id = st.session_state.get("v53_run_id") or str(s.get("Run ID") or "")
        _swap_cache_key = f"v53_swap_history::{_swap_run_id}"
        if _swap_cache_key in st.session_state:
            swap_history = st.session_state[_swap_cache_key]
        else:
            try:
                swap_history = swap_tracking.record_and_annotate_swaps(
                    r["bucket_summary"], run_id=_swap_run_id
                )
            except TypeError:
                # Backward compatibility with older swap_tracking.py.
                swap_history = swap_tracking.record_and_annotate_swaps(r["bucket_summary"])
            st.session_state[_swap_cache_key] = swap_history
        if not swap_history.empty:
            recurring = swap_history[swap_history["Status"].astype(str).str.startswith("RECURRING")]
            if not recurring.empty:
                st.error(f"🔴 {len(recurring)} Store Code swap(s) are RECURRING -- already reported in an earlier run and still unresolved in the source data.")
            else:
                st.error(f"🔴 {len(swap_history)} possible Store Code swap(s) detected this run -- see below.")
            st.dataframe(swap_history,use_container_width=True,hide_index=True)

        # V41 item 10: caption corrected -- V40 item 8 sorts Top 20 strictly
        # by absolute SAR difference across ALL non-matched statuses, not by
        # severity first (the old caption still described the pre-V40
        # GL-AMOUNT-EXCEPTION-only, severity-first ordering).
        st.subheader("🔴 Top 20 Exceptions by absolute SAR exposure")
        st.dataframe(r["top_exceptions"],use_container_width=True,hide_index=True)
    else:
        pos_rows=int(s.get("POS Rows",0)); gl_matched=int(s.get("GL Matched",0))
        amount_exc=int(s.get("GL Amount Exceptions",0)); not_posted=int(s.get("GL Not Posted",0))
        review=int(s.get("Review Required",0)); id_mismatch=int(s.get("Identifier Mismatch",0))
        incomplete=int(s.get("POS Data Incomplete",0)); unmatched_gl=int(s.get("Unmatched GL Rows",0))
        exceptions=amount_exc+not_posted+review+id_mismatch+incomplete
        st.subheader("Control Dashboard")
        a,b,c,d=st.columns(4)
        a.metric("Overall",overall); b.metric("POS Rows",f"{pos_rows:,}"); c.metric("GL Matched",f"{gl_matched:,}"); d.metric("Exceptions",f"{exceptions:,}")

    tab_labels=["All Results","GL Matched","Amount Exceptions"]
    if is_bucket:
        tab_labels += ["Upload Incomplete"]
    tab_labels += ["Not Posted / ID","Unmatched GL"]
    if is_bucket:
        tab_labels += ["Store Date Buckets","Chronic Stores","Duplicate Dates","Swap History"]
    tabs=st.tabs(tab_labels)
    ti = iter(tabs)
    with next(ti): st.dataframe(r["detail"],use_container_width=True,hide_index=True)
    with next(ti): st.dataframe(r["matched"],use_container_width=True,hide_index=True)
    with next(ti): st.dataframe(r["exceptions"][r["exceptions"]["Status"]=="GL AMOUNT EXCEPTION"],use_container_width=True,hide_index=True)
    if is_bucket:
        with next(ti):
            st.caption("GL clearing-account activity with no matching POS/provider evidence for this exact Store+Date, separated out from real accounting exceptions. UPLOAD INCOMPLETE = no such provider was uploaded at all -- likely a missing file. PROVIDER MAPPING REQUIRED (V41) = a provider WAS uploaded but this app doesn't know how to map its value to a GL account -- check the mapping before assuming a file is missing.")
            st.dataframe(r["exceptions"][r["exceptions"]["Status"].isin(["UPLOAD INCOMPLETE","PROVIDER MAPPING REQUIRED"])],use_container_width=True,hide_index=True)
    not_posted_statuses=["GL NOT POSTED","IDENTIFIER MISMATCH","POS DATA INCOMPLETE"]+(["GL REVIEW REQUIRED"] if not is_bucket else [])
    with next(ti): st.dataframe(r["exceptions"][r["exceptions"]["Status"].isin(not_posted_statuses)],use_container_width=True,hide_index=True)
    with next(ti): st.dataframe(r["unmatched_gl"],use_container_width=True,hide_index=True)
    if is_bucket:
        with next(ti):
            st.caption("One row per Store + Date bucket, sorted worst-first within each status, EXTREME/HIGH severity first -- the best view for spotting duplicate files, store-code swaps, or missing-provider coverage gaps.")
            st.dataframe(r["bucket_summary"],use_container_width=True,hide_index=True)
        with next(ti):
            st.caption("Stores failing on nearly every date in this run. Failure Pattern (V41) splits failing days into four causes -- CHRONIC UPLOAD INCOMPLETE, CHRONIC AMOUNT VARIANCE, CHRONIC GL NOT POSTED, CHRONIC UNMATCHED GL -- and names whichever one dominates, or CHRONIC (MIXED CAUSES) with a per-cause breakdown when no single cause dominates.")
            st.dataframe(r.get("chronic_stores"),use_container_width=True,hide_index=True)
        with next(ti):
            st.caption("Dates where many different stores excluded duplicate POS rows -- a likely sign that whole day's POS export was uploaded twice.")
            st.dataframe(r.get("duplicate_dates"),use_container_width=True,hide_index=True)
        with next(ti):
            st.caption("Every Store Code swap detected this run, annotated with its full history across past runs (requires this app's local storage to persist between sessions). Times Seen now counts distinct reconciliation runs (V40) -- rerunning the page for an unrelated widget no longer inflates it.")
            st.dataframe(swap_history,use_container_width=True,hide_index=True)

    b=io.BytesIO()
    with pd.ExcelWriter(b,engine="openpyxl") as w:
        # V42.2: keep the downloadable Summary sheet intentionally small.
        # The wide run/control KPIs remain on the Streamlit dashboard; the
        # Excel Summary contains only the two Finance management views the
        # user asked for: Store-wise and full GL-wise comparison.
        # V42.4: GL-wise now carries the same Status verdict as Store-wise
        # (same tolerance rule); both tables get OK/REVIEW conditional
        # formatting and an explicit Total row so the tie-out that's
        # already proven internally is visible in the workbook itself,
        # not just provable by manually summing the rows.
        _store_summary = r.get("store_summary", pd.DataFrame())
        _gl_summary = r.get("gl_summary", pd.DataFrame())
        # "Mapping Required" drives highlighting only -- it's a raw boolean
        # flag, not something Finance needs to see as its own column, so
        # it's captured here and dropped before writing the visible table.
        _gl_mapping_required = (
            _gl_summary["Mapping Required"].tolist() if "Mapping Required" in _gl_summary.columns else []
        )
        _gl_summary_display = _gl_summary.drop(columns=["Mapping Required"], errors="ignore")
        _store_summary.to_excel(w,index=False,sheet_name="Summary",startrow=2,startcol=0)
        _gl_summary_display.to_excel(w,index=False,sheet_name="Summary",startrow=2,startcol=6)
        _ws = w.sheets["Summary"]
        _ws["A1"] = "STORE-WISE SUMMARY"
        _ws["G1"] = "GL-WISE SUMMARY"

        # Row directly below each table's last data row, for the Total line.
        _store_total_row = 4 + len(_store_summary)
        _gl_total_row = 4 + len(_gl_summary)

        if not _store_summary.empty:
            _ws.cell(row=_store_total_row, column=1, value="TOTAL")
            _ws.cell(row=_store_total_row, column=2, value=float(_store_summary["POS Total"].sum()))
            _ws.cell(row=_store_total_row, column=3, value=float(_store_summary["GL Total"].sum()))
            _ws.cell(row=_store_total_row, column=4, value=float(_store_summary["Difference"].sum()))
        if not _gl_summary.empty:
            _ws.cell(row=_gl_total_row, column=7, value="TOTAL")
            _ws.cell(row=_gl_total_row, column=9, value=float(_gl_summary["GL Total"].sum()))
            _ws.cell(row=_gl_total_row, column=10, value=float(_gl_summary["POS Total"].sum()))
            _ws.cell(row=_gl_total_row, column=11, value=float(_gl_summary["Difference"].sum()))

        try:
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            _title_fill = PatternFill("solid", fgColor="1F4E78")
            _head_fill = PatternFill("solid", fgColor="D9EAF7")
            _ok_fill = PatternFill("solid", fgColor="C6EFCE")
            _review_fill = PatternFill("solid", fgColor="FFC7CE")
            _ok_font = Font(color="006100")
            _review_font = Font(color="9C0006")
            _total_border = Border(top=Side(style="thin"))

            for _cell in (_ws["A1"], _ws["G1"]):
                _cell.font = Font(bold=True, color="FFFFFF")
                _cell.fill = _title_fill

            # Header row now spans 6 columns per table (Status added to
            # GL-wise): A:E (store-wise) and G:L (GL-wise, was G:K).
            for _cell in list(_ws[3])[0:5] + list(_ws[3])[6:12]:
                _cell.font = Font(bold=True)
                _cell.fill = _head_fill
                _cell.alignment = Alignment(horizontal="center")

            for _row in _ws.iter_rows(min_row=4, max_row=max(_store_total_row, _gl_total_row)):
                for _cell in (_row[1:4] + _row[8:11]):
                    if _cell.value is not None and isinstance(_cell.value, (int, float)):
                        _cell.number_format = '#,##0.00;[Red]-#,##0.00'

            # OK/REVIEW conditional formatting: Store-wise Status is column
            # E (index 5), GL-wise Status is column L (index 12) now that
            # it carries the same verdict as Store-wise.
            if not _store_summary.empty:
                for _r in range(4, _store_total_row):
                    _cell = _ws.cell(row=_r, column=5)
                    if _cell.value == "OK":
                        _cell.fill, _cell.font = _ok_fill, _ok_font
                    elif _cell.value == "REVIEW":
                        _cell.fill, _cell.font = _review_fill, _review_font
            if not _gl_summary.empty:
                for _r in range(4, _gl_total_row):
                    _cell = _ws.cell(row=_r, column=12)
                    if _cell.value == "OK":
                        _cell.fill, _cell.font = _ok_fill, _ok_font
                    elif _cell.value == "REVIEW":
                        _cell.fill, _cell.font = _review_fill, _review_font

            # V42.5: rows needing a GL/provider mapping are a categorically
            # different problem than an amount variance -- missing
            # configuration, not an accounting difference -- so they get a
            # distinct highlight (orange) across the whole row, overriding
            # the plain OK/REVIEW coloring set above for those specific rows.
            if not _gl_summary.empty and _gl_mapping_required:
                _mapping_fill = PatternFill("solid", fgColor="FCE4D6")
                _mapping_font = Font(color="974706", bold=True)
                for _offset, _needs_mapping in enumerate(_gl_mapping_required):
                    if not _needs_mapping:
                        continue
                    _r = 4 + _offset
                    for _c in range(7, 13):
                        _cell = _ws.cell(row=_r, column=_c)
                        _cell.fill = _mapping_fill
                        if _c in (7, 8):
                            _cell.font = _mapping_font

            # Bold the Total rows and give them a top border to set them
            # apart from the data above.
            if not _store_summary.empty:
                for _c in range(1, 6):
                    _cell = _ws.cell(row=_store_total_row, column=_c)
                    _cell.font = Font(bold=True)
                    _cell.border = _total_border
                    if isinstance(_cell.value, (int, float)):
                        _cell.number_format = '#,##0.00;[Red]-#,##0.00'
            if not _gl_summary.empty:
                for _c in range(7, 13):
                    _cell = _ws.cell(row=_gl_total_row, column=_c)
                    _cell.font = Font(bold=True)
                    _cell.border = _total_border
                    if isinstance(_cell.value, (int, float)):
                        _cell.number_format = '#,##0.00;[Red]-#,##0.00'

            _ws.column_dimensions["A"].width = 12
            _ws.column_dimensions["B"].width = 16
            _ws.column_dimensions["C"].width = 16
            _ws.column_dimensions["D"].width = 16
            _ws.column_dimensions["E"].width = 11
            _ws.column_dimensions["G"].width = 22
            _ws.column_dimensions["H"].width = 34
            _ws.column_dimensions["I"].width = 16
            _ws.column_dimensions["J"].width = 16
            _ws.column_dimensions["K"].width = 16
            _ws.column_dimensions["L"].width = 11
        except Exception:
            # Formatting must never block the reconciliation download.
            pass
        r["detail"].to_excel(w,index=False,sheet_name="POS to GL")
        r["matched"].to_excel(w,index=False,sheet_name="GL Matched")
        r["exceptions"].to_excel(w,index=False,sheet_name="Exceptions")
        r["unmatched_gl"].to_excel(w,index=False,sheet_name="Unmatched GL")
        if is_bucket:
            r["bucket_summary"].to_excel(w,index=False,sheet_name="Store Date Buckets")
            r["top_exceptions"].to_excel(w,index=False,sheet_name="Top 20 Exceptions")
            r.get("chronic_stores",pd.DataFrame()).to_excel(w,index=False,sheet_name="Chronic Stores")
            r.get("duplicate_dates",pd.DataFrame()).to_excel(w,index=False,sheet_name="Duplicate Dates")
            swap_history.to_excel(w,index=False,sheet_name="Swap History")
        # V41 item 4 / V42 items 2+3: Import Audit ships in the export too,
        # not just shown in the Validate step -- a quarantined file/sheet
        # should be traceable from the delivered workbook itself, not only
        # on screen. Reads from the RESULT's own frozen audit (set at RUN
        # time, above) rather than st.session_state.v53_import_audit -- the
        # latter only ever reflected the last Validate click, which is not
        # necessarily the files that produced THIS result.
        import_audit = r.get("import_audit")
        if import_audit is not None and not import_audit.empty:
            import_audit.to_excel(w,index=False,sheet_name="Import Audit")
    st.download_button("DOWNLOAD POS-GL RECONCILIATION",b.getvalue(),"RetailReconAI_POS_to_GL_Reconciliation.xlsx",mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",use_container_width=True)
else:
    st.subheader("Control Rule")
    st.markdown("### POS Statement Amount ↔ D365 GL Amount")
    st.write("Store Code and Date identify the bucket. Amount alone proves or rejects the match. Merchant ID, Provider, Terminal and Auth Code stay available as investigation detail only.")
