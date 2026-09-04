from __future__ import annotations

import io
from datetime import datetime

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.formatting.rule import CellIsRule
from openpyxl.utils import get_column_letter

from logic.commission_control import norm_payment, validate_commission_transactions


PAYMENTS = ["MADA", "VISA", "MASTER", "AMEX", "TABBY", "TAMARA", "TAP", "FLOOSS", "PAYLATER", "DEEMA"]

TX_COLUMNS = [
    "Store Code", "Date", "Receipt ID", "Auth Code", "D365 Source Type",
    "Reservation Type", "Reservation Flow", "Sales Order", "Customer Account",
    "Customer Name", "Description", "Cash Classification", "Cash Amount", "Reservation Cash", "Order Balance",
    "Reservation Report Total", "Reservation Auth Resolution",
    "D365 MADA", "POS MADA", "Diff MADA",
    "D365 VISA", "POS VISA", "Diff VISA",
    "D365 MASTER", "POS MASTER", "Diff MASTER",
    "D365 AMEX", "POS AMEX", "Diff AMEX",
    "D365 TABBY", "POS TABBY", "Diff TABBY",
    "D365 TAMARA", "POS TAMARA", "Diff TAMARA",
    "D365 TAP", "POS TAP", "Diff TAP",
    "D365 FLOOSS", "POS FLOOSS", "Diff FLOOSS",
    "D365 PAYLATER", "POS PAYLATER", "Diff PAYLATER",
    "D365 DEEMA", "POS DEEMA", "Diff DEEMA",
    "D365 Total", "POS Total", "Total Difference",
    "D365 Tender", "POS Tender", "POS Date", "Posting Date", "Terminal ID",
    "Commission", "VAT", "Net Amount", "Source", "Status", "Remarks"
]


def _master_name(payment):
    p = norm_payment(payment)
    if p == "MASTERCARD":
        return "MASTER"
    return p


def _zero_tx():
    d = {c: "" for c in TX_COLUMNS}
    for p in PAYMENTS:
        d[f"D365 {p}"] = 0.0
        d[f"POS {p}"] = 0.0
        d[f"Diff {p}"] = 0.0

    d.update({
        "Cash Amount": 0.0,
        "Reservation Cash": 0.0,
        "Order Balance": 0.0,
        "Reservation Report Total": 0.0,
        "D365 Total": 0.0,
        "POS Total": 0.0,
        "Total Difference": 0.0,
        "Commission": 0.0,
        "VAT": 0.0,
        "Net Amount": 0.0,
    })
    return d


def transaction_reconciliation(result):
    rows = []

    m = result.get("matched", pd.DataFrame())
    for _, r in m.iterrows():
        x = _zero_tx()
        p = _master_name(r.get("Payment Type"))
        d365 = float(r.get("D365 Amount", 0) or 0)
        pos = float(r.get("POS Amount", 0) or 0)
        x.update({
            "Store Code": r.get("Store Code", ""),
            "Date": r.get("Date", pd.NaT),
            "Receipt ID": r.get("Receipt ID", ""),
            "Auth Code": r.get("Auth Code", ""),
            "D365 Source Type": "Store Tender",
            "D365 Total": d365,
            "POS Total": pos,
            "Total Difference": round(d365 - pos, 2),
            "D365 Tender": p,
            "POS Tender": p,
            "POS Date": r.get("POS Date", pd.NaT),
            "Posting Date": r.get("Posting Date", pd.NaT),
            "Terminal ID": r.get("Terminal ID", ""),
            "Commission": float(r.get("Commission", 0) or 0),
            "VAT": float(r.get("VAT", 0) or 0),
            "Net Amount": float(r.get("Net Amount", 0) or 0),
            "Source": r.get("Source File", ""),
            "Status": "Matched" if str(r.get("Status", "")).lower() == "matched" else "Review",
            "Remarks": "Reconciled successfully." if str(r.get("Status", "")).lower() == "matched"
                       else "Authorization matched but amount requires Finance review.",
        })
        if p in PAYMENTS:
            x[f"D365 {p}"] = d365
            x[f"POS {p}"] = pos
            x[f"Diff {p}"] = round(d365 - pos, 2)
        rows.append(x)

    cash = result.get("cash_transactions", pd.DataFrame())
    for _, r in cash.iterrows():
        x = _zero_tx()
        amt = float(r.get("Cash Amount", r.get("D365 Amount", 0)) or 0)
        cls = str(r.get("Cash Classification", "")).strip() or ("Cash Sales" if amt > 0 else "Cash Refund")
        x.update({
            "Store Code": r.get("Store Code", ""),
            "Date": r.get("Date", pd.NaT),
            "Receipt ID": r.get("Receipt ID", ""),
            "Auth Code": r.get("Auth Code", ""),
            "D365 Source Type": "Store Tender",
            "Cash Classification": cls,
            "Cash Amount": amt,
            "D365 Total": amt,
            "POS Total": 0.0,
            "Total Difference": 0.0,
            "D365 Tender": "CASH",
            "POS Tender": "",
            "Status": cls,
            "Remarks": "Cash transaction from D365 Store Tender. No POS/provider settlement required.",
        })
        rows.append(x)

    us = result.get("unmatched_sales", pd.DataFrame())
    for _, r in us.iterrows():
        x = _zero_tx()
        p = _master_name(r.get("D365 Payment"))
        d365 = float(r.get("D365 Amount", 0) or 0)
        x.update({
            "Store Code": r.get("Store Code", ""),
            "Date": r.get("Date", pd.NaT),
            "Receipt ID": r.get("Receipt ID", ""),
            "Auth Code": r.get("Auth Code", ""),
            "D365 Source Type": "Store Tender",
            "D365 Total": d365,
            "POS Total": 0.0,
            "Total Difference": d365,
            "D365 Tender": p,
            "POS Tender": "",
            "Status": "Missing POS",
            "Remarks": "D365 transaction found but no matching POS settlement. Please check bank/provider settlement.",
        })
        if p in PAYMENTS:
            x[f"D365 {p}"] = d365
            x[f"POS {p}"] = 0.0
            x[f"Diff {p}"] = d365
        rows.append(x)

    up = result.get("unmatched_pos", pd.DataFrame())
    for _, r in up.iterrows():
        x = _zero_tx()
        p = _master_name(r.get("POS Payment"))
        pos = float(r.get("POS Amount", 0) or 0)
        x.update({
            "Store Code": r.get("POS Store", ""),
            "Date": pd.NaT,
            "Receipt ID": "",
            "Auth Code": r.get("Auth Code", ""),
            "D365 Source Type": "",
            "D365 Total": 0.0,
            "POS Total": pos,
            "Total Difference": -pos,
            "D365 Tender": "",
            "POS Tender": p,
            "POS Date": r.get("POS Date", pd.NaT),
            "Posting Date": r.get("Posting Date", pd.NaT),
            "Terminal ID": r.get("Terminal ID", ""),
            "Commission": float(r.get("Commission", 0) or 0),
            "VAT": float(r.get("VAT", 0) or 0),
            "Net Amount": float(r.get("Net Amount", 0) or 0),
            "Source": r.get("Source File", ""),
            "Status": r.get("Exception Status", "Missing D365"),
            "Remarks": r.get(
                "Reason",
                "POS/provider settlement found but no matching D365 Store Tender transaction."
            ),
        })
        if p in PAYMENTS:
            x[f"D365 {p}"] = 0.0
            x[f"POS {p}"] = pos
            x[f"Diff {p}"] = -pos
        rows.append(x)

    out = pd.DataFrame(rows, columns=TX_COLUMNS)
    if out.empty:
        out = pd.DataFrame(columns=TX_COLUMNS)
    return out


def _merchant_mapping_required_store(value):
    """True only when Store Code is genuinely blank/unresolved."""
    if pd.isna(value):
        return True
    return str(value).strip() == ""


def store_summary(tx):
    flag_col = "Merchant Mapping Required"
    visible_cols = (
        ["Store Code"]
        + [f"{z} {p}" for p in PAYMENTS for z in ("D365", "POS", "Diff")]
        + ["D365 Total", "POS Total", "Total Difference", "Exceptions"]
    )

    if tx.empty:
        return pd.DataFrame(columns=visible_cols + [flag_col])

    rows = []
    for store, g in tx.groupby("Store Code", dropna=False):
        mapping_required = _merchant_mapping_required_store(store)
        r = {
            "Store Code": "⚠ MERCHANT MAPPING REQUIRED" if mapping_required else store,
            flag_col: mapping_required,
        }
        for p in PAYMENTS:
            r[f"D365 {p}"] = g[f"D365 {p}"].sum()
            r[f"POS {p}"] = g[f"POS {p}"].sum()
            r[f"Diff {p}"] = g[f"Diff {p}"].sum()

        r["D365 Total"] = g["D365 Total"].sum()
        r["POS Total"] = g["POS Total"].sum()
        r["Total Difference"] = g["Total Difference"].sum()
        r["Exceptions"] = int((g["Status"] != "Matched").sum())
        rows.append(r)

    return pd.DataFrame(rows, columns=visible_cols + [flag_col])


def payment_summary(tx):
    rows = []
    for p in PAYMENTS:
        d365 = float(tx[f"D365 {p}"].sum()) if not tx.empty else 0.0
        pos = float(tx[f"POS {p}"].sum()) if not tx.empty else 0.0
        rows.append({
            "Payment Type": p,
            "D365 Amount": d365,
            "POS Amount": pos,
            "Difference": d365 - pos,
        })
    return pd.DataFrame(rows)


def settlement_commission(tx, commission_rate_master=None):
    """
    V46B Settlement & Commission control.

    Actual commission/VAT remain source values. Expected values come only from
    the editable Commission Rate Master. PROVIDER_ACTUAL rows never receive an
    invented contract expectation.
    """
    flag_col = "Merchant Mapping Required"
    visible_cols = [
        "Store Code",
        "Payment Type",
        "Transactions",
        "Gross Amount",
        "Actual Commission",
        "Expected Commission",
        "Commission Variance",
        "Actual VAT",
        "Expected VAT",
        "VAT Variance",
        "Net Settlement",
        "Control Status",
    ]

    if tx.empty:
        return pd.DataFrame(columns=visible_cols + [flag_col])

    g = tx[(tx["POS Total"] != 0) & (tx["POS Tender"].astype(str).str.upper() != "CASH")].copy()
    if g.empty:
        return pd.DataFrame(columns=visible_cols + [flag_col])

    # Apply the same transaction-level rules as the Commission Validation page
    # before aggregation. This preserves transaction-level rounding.
    control_input = pd.DataFrame({
        "Store Code": g["Store Code"],
        "Payment Type": g["POS Tender"],
        "POS Amount": g["POS Total"],
        "Commission": g["Commission"],
        "VAT": g["VAT"],
        "Net Amount": g["Net Amount"],
        "Auth Code": g["Auth Code"],
    })

    controlled = validate_commission_transactions(
        control_input,
        commission_rate_master if commission_rate_master is not None else pd.DataFrame(),
        payment_col="Payment Type",
        amount_col="POS Amount",
        commission_col="Commission",
        vat_col="VAT",
    )

    controlled["Report Payment Type"] = (
        controlled["Payment Type"]
        .replace({"MASTERCARD": "MASTER"})
    )

    def _aggregate_status(values):
        vals = [str(v) for v in values if str(v).strip()]
        if not vals:
            return "RATE NOT CONFIGURED"

        priority = [
            "SOURCE COMMISSION MISSING",
            "SOURCE VAT MISSING",
            "COMMISSION VARIANCE — REVIEW",
            "VAT VARIANCE",
            "RATE NOT CONFIGURED",
            "CONTRACT RATE PENDING",
            "OK",
        ]
        for status in priority:
            if status in vals:
                return status
        return vals[0]

    rows = []
    for (store, payment), x in controlled.groupby(
        ["Store Code", "Report Payment Type"],
        dropna=False,
    ):
        mapping_required = _merchant_mapping_required_store(store)

        rows.append({
            "Store Code": (
                "⚠ MERCHANT MAPPING REQUIRED"
                if mapping_required
                else store
            ),
            "Payment Type": payment,
            "Transactions": int(len(x)),
            "Gross Amount": round(pd.to_numeric(x["POS Amount"], errors="coerce").fillna(0).sum(), 2),
            "Actual Commission": round(x["Actual Commission"].sum(), 2),
            "Expected Commission": (
                round(x["Expected Commission"].sum(min_count=1), 2)
                if x["Expected Commission"].notna().any()
                else np.nan
            ),
            "Commission Variance": (
                round(x["Commission Variance"].sum(min_count=1), 2)
                if x["Commission Variance"].notna().any()
                else np.nan
            ),
            "Actual VAT": round(x["Actual VAT"].sum(), 2),
            "Expected VAT": (
                round(x["Expected VAT"].sum(min_count=1), 2)
                if x["Expected VAT"].notna().any()
                else np.nan
            ),
            "VAT Variance": (
                round(x["VAT Variance"].sum(min_count=1), 2)
                if x["VAT Variance"].notna().any()
                else np.nan
            ),
            "Net Settlement": round(pd.to_numeric(x["Net Amount"], errors="coerce").fillna(0).sum(), 2),
            "Control Status": _aggregate_status(x["Control Status"].tolist()),
            flag_col: mapping_required,
        })

    return pd.DataFrame(rows, columns=visible_cols + [flag_col])

def settlement_delay(tx):
    m = tx[tx["Status"] == "Matched"].copy()

    if m.empty:
        return pd.DataFrame({
            "Delay Bucket": ["T+0", "T+1", "T+2", "T+3", ">T+3", "Unknown"],
            "Transactions": [0] * 6,
            "% of Matched": [0.0] * 6,
            "D365 Amount": [0.0] * 6,
        })

    sale = pd.to_datetime(m["Date"], errors="coerce")
    pos = pd.to_datetime(m["POS Date"], errors="coerce")
    m["_delay"] = (pos.dt.normalize() - sale.dt.normalize()).dt.days

    def bucket(v):
        if pd.isna(v):
            return "Unknown"
        if v <= 0:
            return "T+0"
        if v == 1:
            return "T+1"
        if v == 2:
            return "T+2"
        if v == 3:
            return "T+3"
        return ">T+3"

    m["_bucket"] = m["_delay"].map(bucket)
    order = ["T+0", "T+1", "T+2", "T+3", ">T+3", "Unknown"]
    rows = []
    total = len(m)

    for b in order:
        x = m[m["_bucket"] == b]
        rows.append({
            "Delay Bucket": b,
            "Transactions": len(x),
            "% of Matched": len(x) / total if total else 0.0,
            "D365 Amount": x["D365 Total"].sum(),
        })

    return pd.DataFrame(rows)


def _write_title(ws, title, subtitle, end_col):
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=end_col)
    ws.cell(1, 1, title)

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=end_col)
    ws.cell(2, 1, subtitle)


def _style_sheet(ws, title_row=1, header_row=4, money_cols=None, percent_cols=None):
    navy = "17365D"
    blue = "1F4E78"
    red = "FCE4D6"
    green = "E2F0D9"
    white = "FFFFFF"
    thin = Side(style="thin", color="D9E1F2")

    ws.sheet_view.showGridLines = False
    ws.freeze_panes = f"A{header_row + 1}"

    for cell in ws[title_row]:
        cell.fill = PatternFill("solid", fgColor=navy)
        cell.font = Font(color=white, bold=True, size=16)
        cell.alignment = Alignment(vertical="center")
    ws.row_dimensions[title_row].height = 25

    for cell in ws[header_row]:
        cell.fill = PatternFill("solid", fgColor=blue)
        cell.font = Font(color=white, bold=True)
        cell.alignment = Alignment(
            horizontal="center",
            vertical="center",
            wrap_text=True,
        )
        cell.border = Border(bottom=thin)
    ws.row_dimensions[header_row].height = 34

    if money_cols:
        for col in money_cols:
            for c in ws[col][header_row:]:
                c.number_format = '#,##0.00'

    if percent_cols:
        for col in percent_cols:
            for c in ws[col][header_row:]:
                c.number_format = '0.00%'

    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        max_len = 0
        for c in col_cells[:200]:
            if c.value is not None:
                max_len = max(max_len, len(str(c.value)))
        ws.column_dimensions[letter].width = min(max(max_len + 2, 11), 34)

    header_map = {
        ws.cell(header_row, c).value: c
        for c in range(1, ws.max_column + 1)
    }

    if "Status" in header_map:
        col = get_column_letter(header_map["Status"])

        ws.conditional_formatting.add(
            f"{col}{header_row + 1}:{col}{ws.max_row}",
            CellIsRule(
                operator="equal",
                formula=['"Matched"'],
                fill=PatternFill("solid", fgColor=green),
            ),
        )

        ws.conditional_formatting.add(
            f"{col}{header_row + 1}:{col}{ws.max_row}",
            CellIsRule(
                operator="notEqual",
                formula=['"Matched"'],
                fill=PatternFill("solid", fgColor=red),
            ),
        )


def _highlight_mapping_required_rows(ws, row_flags, header_row=4):
    """Presentation-only V46A highlight. row_flags aligns with exported data rows."""
    fill = PatternFill("solid", fgColor="FCE4D6")
    dark_orange = "9C5700"

    for offset, required in enumerate(row_flags, start=header_row + 1):
        if not bool(required):
            continue

        for col in range(1, ws.max_column + 1):
            ws.cell(offset, col).fill = fill

        ws.cell(offset, 1).font = Font(
            bold=True,
            color=dark_orange,
        )


def create_reconciliation_pack(result, tolerance=1.0):
    tx = transaction_reconciliation(result)
    exceptions = tx[tx["Status"] != "Matched"].copy()
    store = store_summary(tx)
    payment = payment_summary(tx)
    # V46B: load the same editable Commission Rate Master used by
    # Commission Validation. Export remains available even if DB access fails.
    try:
        import db
        commission_rate_master = db.load_commission_rate_master().copy()
    except Exception:
        commission_rate_master = pd.DataFrame()

    settlement = settlement_commission(
        tx,
        commission_rate_master=commission_rate_master,
    )
    delay = settlement_delay(tx)

    # V46A: keep mapping-required identity internally for formatting only.
    # Never expose the raw boolean flag in the visible Excel workbook.
    mapping_flag = "Merchant Mapping Required"

    store_mapping_flags = (
        store[mapping_flag].tolist()
        if mapping_flag in store.columns
        else [False] * len(store)
    )
    settlement_mapping_flags = (
        settlement[mapping_flag].tolist()
        if mapping_flag in settlement.columns
        else [False] * len(settlement)
    )

    store = store.drop(columns=[mapping_flag], errors="ignore")
    settlement = settlement.drop(columns=[mapping_flag], errors="ignore")

    total = len(tx)
    matched = int((tx["Status"] == "Matched").sum()) if total else 0
    exc = total - matched
    d365 = float(tx["D365 Total"].sum()) if total else 0.0
    pos = float(tx["POS Total"].sum()) if total else 0.0

    now = datetime.now()
    stamp = now.strftime("%d-%b-%Y %H:%M")
    out = io.BytesIO()

    with pd.ExcelWriter(out, engine="openpyxl") as writer:
        # Dashboard
        dash = pd.DataFrame([
            ["Total Transactions", total],
            ["Matched", matched],
            ["Exceptions", exc],
            ["Match %", matched / total if total else 0],
            ["Total D365 Amount", d365],
            ["Total POS/Provider Amount", pos],
            ["Total Difference", d365 - pos],
        ], columns=["Metric", "Value"])

        dash.to_excel(
            writer,
            sheet_name="Dashboard",
            index=False,
            startrow=3,
        )
        ws = writer.book["Dashboard"]
        _write_title(
            ws,
            "RetailRecon AI — Reconciliation Dashboard",
            f"Generated {stamp} | Tolerance SAR {tolerance:.2f}",
            6,
        )
        _style_sheet(ws, header_row=4)
        ws["B8"].number_format = "0.00%"
        for r in range(9, 12):
            ws[f"B{r}"].number_format = '#,##0.00'

        # Main transaction
        tx.to_excel(
            writer,
            sheet_name="Transaction_Reconciliation",
            index=False,
            startrow=3,
        )
        ws = writer.book["Transaction_Reconciliation"]
        _write_title(
            ws,
            "Transaction Reconciliation",
            f"Generated {stamp} | {len(tx):,} row(s)",
            len(TX_COLUMNS),
        )
        _style_sheet(ws, header_row=4)

        for col in range(16, 58):
            letter = get_column_letter(col)
            for c in ws[letter][4:]:
                c.number_format = '#,##0.00'

        for dc in ["B", "AY", "AZ"]:
            for c in ws[dc][4:]:
                c.number_format = 'dd-mmm-yyyy'

        # Exceptions
        exceptions.to_excel(
            writer,
            sheet_name="Exceptions",
            index=False,
            startrow=3,
        )
        ws = writer.book["Exceptions"]
        _write_title(
            ws,
            "Exceptions",
            f"Generated {stamp} | {len(exceptions):,} exception row(s)",
            len(TX_COLUMNS),
        )
        _style_sheet(ws, header_row=4)

        for col in range(16, 58):
            letter = get_column_letter(col)
            for c in ws[letter][4:]:
                c.number_format = '#,##0.00'

        # Store summary
        store.to_excel(
            writer,
            sheet_name="Store_Summary",
            index=False,
            startrow=3,
        )
        ws = writer.book["Store_Summary"]
        _write_title(
            ws,
            "Store Summary",
            f"Generated {stamp}",
            len(store.columns),
        )
        _style_sheet(ws, header_row=4)

        _highlight_mapping_required_rows(
            ws,
            store_mapping_flags,
            header_row=4,
        )

        for col in range(2, ws.max_column):
            letter = get_column_letter(col)
            for c in ws[letter][4:]:
                c.number_format = '#,##0.00'

        # Payment summary
        payment.to_excel(
            writer,
            sheet_name="Payment_Summary",
            index=False,
            startrow=3,
        )
        ws = writer.book["Payment_Summary"]
        _write_title(
            ws,
            "Payment Summary",
            f"Generated {stamp}",
            4,
        )
        _style_sheet(ws, header_row=4)

        for col in ["B", "C", "D"]:
            for c in ws[col][4:]:
                c.number_format = '#,##0.00'

        # Settlement commission
        settlement.to_excel(
            writer,
            sheet_name="Settlement_Commission",
            index=False,
            startrow=3,
        )
        ws = writer.book["Settlement_Commission"]
        _write_title(
            ws,
            "Settlement & Commission — V46B Commission & VAT Control",
            f"Generated {stamp}",
            len(settlement.columns),
        )
        _style_sheet(ws, header_row=4)

        for col in ["D", "E", "F", "G", "H", "I", "J", "K"]:
            for c in ws[col][4:]:
                c.number_format = '#,##0.00'

        # V46B control-status highlight; V46A merchant mapping orange remains authoritative
        # because mapping-required rows are re-applied afterward.
        status_fill = PatternFill("solid", fgColor="FFF2CC")
        status_col = 12
        for row in range(5, ws.max_row + 1):
            status = str(ws.cell(row, status_col).value or "").strip().upper()
            if status and status != "OK":
                ws.cell(row, status_col).fill = status_fill
                ws.cell(row, status_col).font = Font(bold=True)

        _highlight_mapping_required_rows(
            ws,
            settlement_mapping_flags,
            header_row=4,
        )

        # Settlement delay
        delay.to_excel(
            writer,
            sheet_name="Settlement_Delay",
            index=False,
            startrow=4,
        )
        ws = writer.book["Settlement_Delay"]
        _write_title(
            ws,
            "Settlement Delay Analysis",
            (
                f"Generated {stamp} | Matched transactions only | "
                "Delay = POS/provider date - D365 sale date"
            ),
            10,
        )
        ws["A4"] = "Delay Distribution — Matched Transactions Only"
        ws["A4"].font = Font(bold=True, size=12)
        _style_sheet(ws, header_row=5)

        for c in ws["C"][5:]:
            c.number_format = "0.00%"

        for c in ws["D"][5:]:
            c.number_format = '#,##0.00'

    out.seek(0)
    return out.getvalue()
